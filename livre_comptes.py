#!/usr/bin/env python3
"""
Génère un livre de comptes Excel à partir de relevés bancaires PDF.

- Si le fichier de sortie n'existe pas : il est créé avec un onglet pour l'année.
- Si le fichier existe : un nouvel onglet est ajouté pour l'année traitée.
- Un onglet de synthèse pluriannuelle est automatiquement mis à jour.

Usage:
    python livre_comptes.py releve1.pdf releve2.pdf [...] [-o sortie.xlsx] [-a ANNEE] [-f]
"""

import argparse
import re
import sys
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.worksheet.datavalidation import DataValidation

try:
    import camelot
    import pandas as pd
except ImportError:
    print("❌ Installez les dépendances : pip install camelot-py[cv] pandas openpyxl")
    sys.exit(1)


# ============================================================
# STYLES
# ============================================================
COULEURS = {
    "titre":        "1F3864",   # bleu marine
    "recette":      "1E8449",   # vert foncé
    "recette_clair":"D5F5E3",   # vert clair
    "depense":      "922B21",   # rouge foncé
    "depense_clair":"FADBD8",   # rouge clair
    "recap":        "1A5276",   # bleu foncé
    "recap_clair":  "D6EAF8",   # bleu clair
    "synthese":     "4A235A",   # violet foncé
    "synthese_clair":"E8DAEF",  # violet clair
    "solde_pos":    "1E8449",   # vert
    "solde_neg":    "922B21",   # rouge
    "alternance":   "F2F3F4",   # gris très clair
    "blanc":        "FFFFFF",
    "or":           "D4AC0D",
}

def style_bordure(fin=False):
    s = "thin" if not fin else "hair"
    b = Side(style=s)
    return Border(left=b, right=b, top=b, bottom=b)

def fill(couleur):
    return PatternFill("solid", fgColor=couleur)

def police(gras=False, taille=11, couleur="000000", italique=False):
    return Font(bold=gras, size=taille, color=couleur, italic=italique)

def centrer(horizontal="center", vertical="center", wrap=False):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


# ============================================================
# CATÉGORIES
# ============================================================
CATEGORIES_RECETTES = [
    "Cotisations", "Buvette/Tournoi", "Sponsoring/Dons",
    "Subventions", "Autres", "À préciser"
]
CATEGORIES_DEPENSES = [
    "Arbitrage", "Licences/FFBB", "Matériel sportif", "Assurance",
    "Frais bancaires", "Buvette/Tournoi", "Location salle",
    "Déplacements", "Autres", "À préciser"
]

REGLES_RECETTES = [
    (r"cotisation|adhesion|licence", "Cotisations"),
    (r"rem chq|remise.*chq|remise de cheque", "Buvette/Tournoi"),
    (r"regul.*verst|regul.*gab",    "Buvette/Tournoi"),
    (r"loomis|gab|versement esp",   "Buvette/Tournoi"),
    (r"tournoi|buvette|snack",      "Buvette/Tournoi"),
    (r"sponsor|parrain|don",        "Sponsoring/Dons"),
    (r"subvention|aide|grant",      "Subventions"),
    (r"virement.*recu|vir.*recu",   "Sponsoring/Dons"),
]
REGLES_DEPENSES = [
    (r"retrait|gab|distributeur",                   "Buvette/Tournoi"),
    (r"cheque|chq(?!.*rem)",                        "Autres"),
    (r"smacl|maif|assurance|insurance",             "Assurance"),
    (r"intérêts|interets|commission|frais bancaires|cotisation bancaire", "Frais bancaires"),
    (r"arbitre",                                    "Arbitrage"),
    (r"ffbb|licence|federation",                    "Licences/FFBB"),
    (r"décathlon|decathlon|intersport|sport 2000|go sport|matériel", "Matériel sportif"),
    (r"location|salle|gymnase|terrain",             "Location salle"),
    (r"transport|essence|carburant|péage|deplacement", "Déplacements"),
    (r"tournoi|buvette|restauration",               "Buvette/Tournoi"),
]


def categoriser(libelle: str, regles: list, defaut="À préciser") -> str:
    for pattern, cat in regles:
        if re.search(pattern, libelle.lower()):
            return cat
    return defaut


# ============================================================
# EXTRACTION PDF
# ============================================================


def parse_montant(montant_str: str) -> float:
    """Parse un montant depuis une chaîne, gère différents formats."""
    if not montant_str or montant_str.strip() == "":
        return 0.0
    
    # Nettoyer la chaîne
    montant_clean = montant_str.strip().replace(" ", "").replace(",", ".")
    
    try:
        return float(montant_clean)
    except ValueError:
        # Tenter de nettoyer davantage
        import string
        # Garder seulement chiffres, point, virgule et signe moins/plus
        allowed = string.digits + ".,-+"
        montant_clean = "".join(c for c in montant_clean if c in allowed)
        montant_clean = montant_clean.replace(",", ".")
        
        try:
            return float(montant_clean)
        except ValueError:
            return 0.0


# Fonctions supprimées car remplacées par l'extraction avec camelot
# - classifier_operation() : remplacée par la lecture directe des colonnes débit/crédit
# - analyser_ligne_operation() : remplacée par le traitement pandas des DataFrames


def extraire_operations_pdf_camelot(pdf_path: Path, annee: int) -> tuple[list, list]:
    """Extrait les opérations depuis un PDF en utilisant camelot et pandas."""
    recettes, depenses = [], []

    # Essayer différents paramètres camelot pour maximiser la compatibilité
    paramètres_camelot = [
        {"flavor": "stream", "pages": "all", "edge_tol": 500},  # Pour relevés 2023+
        {"flavor": "stream", "pages": "all"},                   # Pour relevés 2019
        {"flavor": "stream", "pages": "all", "row_tol": 10},    # Paramètres alternatifs
    ]
    
    tables = None
    param_utilisé = None
    
    for i, params in enumerate(paramètres_camelot):
        try:
            print(f"    🔧 Test paramètre {i+1}: {params}")
            tables_test = camelot.read_pdf(str(pdf_path), strip_text="\n", **params)
            
            # Vérifier si on trouve des tables avec des opérations
            tables_avec_operations = []
            for table in tables_test:
                df = table.df
                # Une table d'opérations doit avoir au moins 5 colonnes et des dates
                if df.shape[1] >= 5:
                    for idx, row in df.iterrows():
                        if idx < 4:
                            continue
                        date_col = str(row.iloc[0]).strip()
                        if re.match(r'\d{2}\.\d{2}', date_col):
                            tables_avec_operations.append(table)
                            break
            
            if tables_avec_operations:
                tables = tables_avec_operations
                param_utilisé = params
                print(f"    ✅ Succès avec paramètre {i+1}: {len(tables)} table(s) avec opérations")
                break
            elif tables_test:
                print(f"    ⚠️  Paramètre {i+1}: {len(tables_test)} table(s) trouvée(s) mais sans opérations détectées")
            else:
                print(f"    ❌ Paramètre {i+1}: aucune table trouvée")
                
        except Exception as e:
            print(f"    ❌ Erreur avec paramètre {i+1}: {e}")
            continue
    
    if not tables:
        print(f"❌ Aucune table d'opérations trouvée dans {pdf_path.name}")
        return [], []
        
    print(f"    📊 {len(tables)} table(s) d'opérations détectée(s)")

    operations_trouvees = 0
    
    # Traiter chaque table
    for i, table in enumerate(tables):
        df = table.df
        print(f"    📋 Table {i+1}: {len(df)} lignes")
        
        # Parcourir chaque ligne du DataFrame
        for idx, row in df.iterrows():
            # Ignorer les en-têtes et lignes vides
            if idx < 2:  # Les 2 premières lignes sont généralement les en-têtes
                continue
                
            # Vérifier si c'est une ligne d'opération (avec dates)
            date_ope = str(row.iloc[0]).strip() if len(row) > 0 else ""
            date_val = str(row.iloc[1]).strip() if len(row) > 1 else ""
            libelle = str(row.iloc[2]).strip() if len(row) > 2 else ""
            
            # Ignorer les lignes sans dates valides
            if not re.match(r'\d{2}\.\d{2}', date_ope):
                continue
                
            # Stratégie d'extraction des montants adaptée aux différents formats
            montant = 0.0
            est_recette = False
            
            # Méthode 1: Vérifier les colonnes débit/crédit traditionnelles (avant-dernière/dernière)
            debit_str = str(row.iloc[-2]).strip() if len(row) >= 2 else ""
            credit_str = str(row.iloc[-1]).strip() if len(row) >= 1 else ""
            
            # Nettoyer les montants (enlever les caractères parasites)
            debit_str = re.sub(r'[¨\s]', '', debit_str)
            credit_str = re.sub(r'[¨\s]', '', credit_str)
            
            if debit_str and debit_str != "nan" and debit_str != "":
                # C'est un débit (dépense)
                montant = parse_montant(debit_str)
                est_recette = False
            elif credit_str and credit_str != "nan" and credit_str != "":
                # C'est un crédit (recette)  
                montant = parse_montant(credit_str)
                est_recette = True
            else:
                # Méthode 2: Chercher un montant dans toutes les colonnes (pour format 2023)
                for col_idx in range(3, len(row)):  # Commencer après le libellé
                    val_str = str(row.iloc[col_idx]).strip()
                    val_clean = re.sub(r'[¨\s]', '', val_str)
                    
                    if val_clean and val_clean != "nan" and val_clean != "":
                        # Vérifier si c'est un montant valide (contient chiffres et virgule)
                        if re.match(r'^\d{1,4}(?:\s?\d{3})*,\d{2}$', val_clean):
                            montant = parse_montant(val_clean)
                            if montant > 0:
                                # Déterminer si c'est débit ou crédit selon le libellé
                                libelle_upper = libelle.upper()
                                if any(mot in libelle_upper for mot in ["REM CHQ", "REMISE", "VERSEMENT", "DEPOT", "VIREMENT"]):
                                    est_recette = True
                                else:
                                    est_recette = False
                                break
                
                # Si aucun montant trouvé, ignorer cette ligne
                if montant <= 0:
                    continue
                
            # Ignorer les montants nuls ou invalides
            if montant <= 0:
                continue
                
            operations_trouvees += 1
            
            # Convertir format de date DD.MM en DD/MM
            date_ope_formatted = date_ope.replace('.', '/')
            date_val_formatted = date_val.replace('.', '/')
            
            # Créer l'opération
            op = [
                f"{date_ope_formatted}/{annee}",
                f"{date_val_formatted}/{annee}",
                libelle.strip(),
                "",   # référence
                montant,
                "",   # catégorie (sera remplie ci-dessous)
                pdf_path.stem,
            ]

            if est_recette:
                op[5] = categoriser(libelle, REGLES_RECETTES)
                recettes.append(op)
            else:
                op[5] = categoriser(libelle, REGLES_DEPENSES)
                depenses.append(op)

    print(f"    📋 {operations_trouvees} opérations extraites")
    return recettes, depenses


# ============================================================
# ÉCRITURE D'UN ONGLET ANNUEL
# ============================================================
def ecrire_onglet_annuel(ws, recettes: list, depenses: list, annee: int):
    """Remplit un onglet avec recettes, dépenses, récap et ventilation."""

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Largeurs de colonnes
    largeurs = {"A": 14, "B": 14, "C": 45, "D": 18, "E": 14, "F": 22, "G": 20}
    for col, larg in largeurs.items():
        ws.column_dimensions[col].width = larg

    entetes = ["Date opération", "Date valeur", "Libellé", "Référence",
               "Montant (€)", "Catégorie", "Source"]

    ligne = 1

    # ─── TITRE GÉNÉRAL ──────────────────────────────────────────────────────
    ws.merge_cells(f"A{ligne}:G{ligne}")
    c = ws.cell(ligne, 1, f"📒  LIVRE DE COMPTES {annee}")
    c.font      = police(gras=True, taille=16, couleur="FFFFFF")
    c.fill      = fill(COULEURS["titre"])
    c.alignment = centrer()
    c.border    = style_bordure()
    ws.row_dimensions[ligne].height = 32
    ligne += 2


    # ═══════════════════════════════════════════════════════════════════════
    # SECTION RECETTES
    # ═══════════════════════════════════════════════════════════════════════
    ws.merge_cells(f"A{ligne}:G{ligne}")
    c = ws.cell(ligne, 1, "💰  RECETTES")
    c.font      = police(gras=True, taille=13, couleur="FFFFFF")
    c.fill      = fill(COULEURS["recette"])
    c.alignment = centrer()
    c.border    = style_bordure()
    ws.row_dimensions[ligne].height = 26
    ligne += 1

    # En-têtes recettes
    ligne_entete_rec = ligne
    for col, titre in enumerate(entetes, 1):
        c = ws.cell(ligne, col, titre)
        c.font      = police(gras=True, couleur="FFFFFF")
        c.fill      = fill(COULEURS["recette"])
        c.alignment = centrer()
        c.border    = style_bordure()
    ws.row_dimensions[ligne].height = 20
    ligne += 1

    # Données recettes
    ligne_debut_rec = ligne
    cats_rec_dv = ",".join(CATEGORIES_RECETTES)
    dv_rec = DataValidation(type="list", formula1=f'"{cats_rec_dv}"', showDropDown=False)
    ws.add_data_validation(dv_rec)

    for i, op in enumerate(recettes):
        bg = COULEURS["blanc"] if i % 2 == 0 else COULEURS["alternance"]
        for col, val in enumerate(op, 1):
            c = ws.cell(ligne, col, val)
            c.fill      = fill(bg)
            c.border    = style_bordure(fin=True)
            c.alignment = centrer("left") if col == 3 else centrer()
            if col == 5:
                c.number_format = '#,##0.00 €'
                c.font = police(couleur=COULEURS["recette"])
        dv_rec.add(ws.cell(ligne, 6))
        ligne += 1

    ligne_fin_rec = ligne - 1

    # Total recettes
    ws.merge_cells(f"A{ligne}:D{ligne}")
    c = ws.cell(ligne, 1, "TOTAL RECETTES")
    c.font      = police(gras=True, couleur="FFFFFF")
    c.fill      = fill(COULEURS["recette"])
    c.alignment = centrer("right")
    c.border    = style_bordure()

    col_total_rec = 5
    cell_total_rec = ws.cell(ligne, col_total_rec)
    if recettes:
        cell_total_rec.value = f"=SUM(E{ligne_debut_rec}:E{ligne_fin_rec})"
    else:
        cell_total_rec.value = 0
    cell_total_rec.font         = police(gras=True, couleur="FFFFFF")
    cell_total_rec.fill         = fill(COULEURS["recette"])
    cell_total_rec.number_format = '#,##0.00 €'
    cell_total_rec.alignment    = centrer()
    cell_total_rec.border       = style_bordure()
    ref_total_rec = f"E{ligne}"   # pour récap
    ws.row_dimensions[ligne].height = 22
    ligne += 2


    # ═══════════════════════════════════════════════════════════════════════
    # SECTION DÉPENSES
    # ═══════════════════════════════════════════════════════════════════════
    ws.merge_cells(f"A{ligne}:G{ligne}")
    c = ws.cell(ligne, 1, "💸  DÉPENSES")
    c.font      = police(gras=True, taille=13, couleur="FFFFFF")
    c.fill      = fill(COULEURS["depense"])
    c.alignment = centrer()
    c.border    = style_bordure()
    ws.row_dimensions[ligne].height = 26
    ligne += 1

    # En-têtes dépenses
    for col, titre in enumerate(entetes, 1):
        c = ws.cell(ligne, col, titre)
        c.font      = police(gras=True, couleur="FFFFFF")
        c.fill      = fill(COULEURS["depense"])
        c.alignment = centrer()
        c.border    = style_bordure()
    ws.row_dimensions[ligne].height = 20
    ligne += 1

    # Données dépenses
    ligne_debut_dep = ligne
    cats_dep_dv = ",".join(CATEGORIES_DEPENSES)
    dv_dep = DataValidation(type="list", formula1=f'"{cats_dep_dv}"', showDropDown=False)
    ws.add_data_validation(dv_dep)

    for i, op in enumerate(depenses):
        bg = COULEURS["blanc"] if i % 2 == 0 else COULEURS["alternance"]
        for col, val in enumerate(op, 1):
            c = ws.cell(ligne, col, val)
            c.fill      = fill(bg)
            c.border    = style_bordure(fin=True)
            c.alignment = centrer("left") if col == 3 else centrer()
            if col == 5:
                c.number_format = '#,##0.00 €'
                c.font = police(couleur=COULEURS["depense"])
        dv_dep.add(ws.cell(ligne, 6))
        ligne += 1

    ligne_fin_dep = ligne - 1

    # Total dépenses
    ws.merge_cells(f"A{ligne}:D{ligne}")
    c = ws.cell(ligne, 1, "TOTAL DÉPENSES")
    c.font      = police(gras=True, couleur="FFFFFF")
    c.fill      = fill(COULEURS["depense"])
    c.alignment = centrer("right")
    c.border    = style_bordure()

    cell_total_dep = ws.cell(ligne, 5)
    if depenses:
        cell_total_dep.value = f"=SUM(E{ligne_debut_dep}:E{ligne_fin_dep})"
    else:
        cell_total_dep.value = 0
    cell_total_dep.font         = police(gras=True, couleur="FFFFFF")
    cell_total_dep.fill         = fill(COULEURS["depense"])
    cell_total_dep.number_format = '#,##0.00 €'
    cell_total_dep.alignment    = centrer()
    cell_total_dep.border       = style_bordure()
    ref_total_dep = f"E{ligne}"
    ws.row_dimensions[ligne].height = 22
    ligne += 2


    # ═══════════════════════════════════════════════════════════════════════
    # RÉCAPITULATIF ANNUEL
    # ═══════════════════════════════════════════════════════════════════════
    ws.merge_cells(f"A{ligne}:G{ligne}")
    c = ws.cell(ligne, 1, "📊  RÉCAPITULATIF ANNUEL")
    c.font      = police(gras=True, taille=13, couleur="FFFFFF")
    c.fill      = fill(COULEURS["recap"])
    c.alignment = centrer()
    c.border    = style_bordure()
    ws.row_dimensions[ligne].height = 26
    ligne += 1

    recap_items = [
        ("Total recettes",   f"={ref_total_rec}", COULEURS["recette_clair"], COULEURS["recette"]),
        ("Total dépenses",   f"={ref_total_dep}", COULEURS["depense_clair"], COULEURS["depense"]),
        ("Solde de l'année", f"={ref_total_rec}-{ref_total_dep}", COULEURS["recap_clair"], COULEURS["recap"]),
    ]

    for label, formule, bg_clair, bg_fort in recap_items:
        ws.merge_cells(f"A{ligne}:D{ligne}")
        c = ws.cell(ligne, 1, label)
        c.font      = police(gras=True, couleur="FFFFFF")
        c.fill      = fill(bg_fort)
        c.alignment = centrer("right")
        c.border    = style_bordure()

        c2 = ws.cell(ligne, 5, formule)
        c2.font         = police(gras=True)
        c2.fill         = fill(bg_clair)
        c2.number_format = '#,##0.00 €'
        c2.alignment    = centrer()
        c2.border       = style_bordure()
        ws.row_dimensions[ligne].height = 22
        ligne += 1

    ligne += 1


    # ═══════════════════════════════════════════════════════════════════════
    # VENTILATION PAR CATÉGORIE
    # ═══════════════════════════════════════════════════════════════════════
    ws.merge_cells(f"A{ligne}:G{ligne}")
    c = ws.cell(ligne, 1, "🎯  VENTILATION PAR CATÉGORIE")
    c.font      = police(gras=True, taille=13, couleur="FFFFFF")
    c.fill      = fill(COULEURS["recap"])
    c.alignment = centrer()
    c.border    = style_bordure()
    ws.row_dimensions[ligne].height = 26
    ligne += 1

    # Recettes par catégorie
    ws.merge_cells(f"A{ligne}:C{ligne}")
    c = ws.cell(ligne, 1, "Catégorie recette")
    c.font = police(gras=True, couleur="FFFFFF")
    c.fill = fill(COULEURS["recette"])
    c.alignment = centrer()
    c.border = style_bordure()

    ws.merge_cells(f"D{ligne}:E{ligne}")
    c = ws.cell(ligne, 4, "Montant (€)")
    c.font = police(gras=True, couleur="FFFFFF")
    c.fill = fill(COULEURS["recette"])
    c.alignment = centrer()
    c.border = style_bordure()
    ligne += 1

    for cat in CATEGORIES_RECETTES:
        if recettes:
            formule = (f'=SUMIF(F{ligne_debut_rec}:F{ligne_fin_rec},'
                       f'"{cat}",E{ligne_debut_rec}:E{ligne_fin_rec})')
        else:
            formule = 0
        ws.merge_cells(f"A{ligne}:C{ligne}")
        c = ws.cell(ligne, 1, cat)
        c.fill = fill(COULEURS["recette_clair"])
        c.border = style_bordure(fin=True)
        c.alignment = centrer("left")

        ws.merge_cells(f"D{ligne}:E{ligne}")
        c2 = ws.cell(ligne, 4, formule)
        c2.number_format = '#,##0.00 €'
        c2.fill = fill(COULEURS["recette_clair"])
        c2.border = style_bordure(fin=True)
        c2.alignment = centrer()
        ligne += 1

    ligne += 1

    # Dépenses par catégorie
    ws.merge_cells(f"A{ligne}:C{ligne}")
    c = ws.cell(ligne, 1, "Catégorie dépense")
    c.font = police(gras=True, couleur="FFFFFF")
    c.fill = fill(COULEURS["depense"])
    c.alignment = centrer()
    c.border = style_bordure()

    ws.merge_cells(f"D{ligne}:E{ligne}")
    c = ws.cell(ligne, 4, "Montant (€)")
    c.font = police(gras=True, couleur="FFFFFF")
    c.fill = fill(COULEURS["depense"])
    c.alignment = centrer()
    c.border = style_bordure()
    ligne += 1

    for cat in CATEGORIES_DEPENSES:
        if depenses:
            formule = (f'=SUMIF(F{ligne_debut_dep}:F{ligne_fin_dep},'
                       f'"{cat}",E{ligne_debut_dep}:E{ligne_fin_dep})')
        else:
            formule = 0
        ws.merge_cells(f"A{ligne}:C{ligne}")
        c = ws.cell(ligne, 1, cat)
        c.fill = fill(COULEURS["depense_clair"])
        c.border = style_bordure(fin=True)
        c.alignment = centrer("left")

        ws.merge_cells(f"D{ligne}:E{ligne}")
        c2 = ws.cell(ligne, 4, formule)
        c2.number_format = '#,##0.00 €'
        c2.fill = fill(COULEURS["depense_clair"])
        c2.border = style_bordure(fin=True)
        c2.alignment = centrer()
        ligne += 1


# ============================================================
# ONGLET SYNTHÈSE PLURIANNUELLE
# ============================================================
def ecrire_onglet_synthese(wb: openpyxl.Workbook):
    """
    Recrée l'onglet '📈 Synthèse' en lisant les totaux de chaque onglet annuel.
    Les totaux sont lus via des références directes aux cellules des onglets annuels.
    """

    # Supprimer l'ancien onglet synthèse s'il existe
    NOM_SYNTHESE = "📈 Synthèse"
    if NOM_SYNTHESE in wb.sheetnames:
        del wb[NOM_SYNTHESE]

    ws = wb.create_sheet(NOM_SYNTHESE, 0)   # en premier
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    # Récupérer les onglets annuels triés
    onglets_annuels = sorted(
        [n for n in wb.sheetnames if n.startswith("Comptes ")],
        key=lambda n: int(n.split()[-1])
    )
    annees = [int(n.split()[-1]) for n in onglets_annuels]

    # Pour chaque onglet annuel, on doit retrouver les cellules
    # contenant le total recettes, total dépenses et solde.
    # On les a nommées via ref_total_rec / ref_total_dep dans ecrire_onglet_annuel.
    # Mais comme on n'a pas stocké les numéros de ligne, on va les chercher.
    def trouver_cellules_recap(ws_annuel):
        """Retourne (cell_total_rec, cell_total_dep) en cherchant les libellés."""
        cell_rec = cell_dep = None
        for row in ws_annuel.iter_rows():
            for cell in row:
                if cell.value == "TOTAL RECETTES":
                    cell_rec = ws_annuel.cell(cell.row, 5)
                if cell.value == "TOTAL DÉPENSES":
                    cell_dep = ws_annuel.cell(cell.row, 5)
        return cell_rec, cell_dep

    # Largeurs
    ws.column_dimensions["A"].width = 10   # Année
    ws.column_dimensions["B"].width = 20   # Total recettes
    ws.column_dimensions["C"].width = 20   # Total dépenses
    ws.column_dimensions["D"].width = 20   # Solde année
    ws.column_dimensions["E"].width = 22   # Solde cumulé
    ws.column_dimensions["F"].width = 16   # Évolution

    ligne = 1

    # ─── TITRE ──────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{ligne}:F{ligne}")
    c = ws.cell(ligne, 1, "📈  SYNTHÈSE PLURIANNUELLE")
    c.font      = police(gras=True, taille=16, couleur="FFFFFF")
    c.fill      = fill(COULEURS["synthese"])
    c.alignment = centrer()
    c.border    = style_bordure()
    ws.row_dimensions[ligne].height = 34
    ligne += 1

    # Sous-titre date de mise à jour
    ws.merge_cells(f"A{ligne}:F{ligne}")
    c = ws.cell(ligne, 1,
                f"Mis à jour le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
    c.font      = police(italique=True, taille=10, couleur="FFFFFF")
    c.fill      = fill(COULEURS["synthese"])
    c.alignment = centrer()
    ws.row_dimensions[ligne].height = 16
    ligne += 1

    # ─── EN-TÊTES ────────────────────────────────────────────────────────────
    entetes_syn = ["Année", "Total recettes", "Total dépenses",
                   "Solde annuel", "Solde cumulé", "Évolution"]
    couleurs_ent = [
        COULEURS["synthese"], COULEURS["recette"], COULEURS["depense"],
        COULEURS["recap"],    COULEURS["recap"],   COULEURS["titre"],
    ]
    for col, (titre, coul) in enumerate(zip(entetes_syn, couleurs_ent), 1):
        c = ws.cell(ligne, col, titre)
        c.font      = police(gras=True, couleur="FFFFFF")
        c.fill      = fill(coul)
        c.alignment = centrer()
        c.border    = style_bordure()
    ws.row_dimensions[ligne].height = 22
    ligne += 1

    # ─── DONNÉES PAR ANNÉE ───────────────────────────────────────────────────
    ligne_debut_donnees = ligne
    lignes_data = {}   # annee -> numéro de ligne dans la synthèse

    for idx, (annee, nom_onglet) in enumerate(zip(annees, onglets_annuels)):
        ws_ann = wb[nom_onglet]
        cell_rec, cell_dep = trouver_cellules_recap(ws_ann)

        nom_safe = nom_onglet.replace("'", "''")  # échapper les apostrophes
        if cell_rec and cell_dep:
            ref_rec = f"'{nom_safe}'!E{cell_rec.row}"
            ref_dep = f"'{nom_safe}'!E{cell_dep.row}"
        else:
            ref_rec = ref_dep = "0"

        bg = COULEURS["blanc"] if idx % 2 == 0 else COULEURS["alternance"]

        # Col A – Année
        c = ws.cell(ligne, 1, annee)
        c.font      = police(gras=True)
        c.fill      = fill(COULEURS["synthese_clair"])
        c.alignment = centrer()
        c.border    = style_bordure()

        # Col B – Total recettes
        c = ws.cell(ligne, 2, f"={ref_rec}")
        c.number_format = '#,##0.00 €'
        c.font      = police(couleur=COULEURS["recette"])
        c.fill      = fill(bg)
        c.alignment = centrer()
        c.border    = style_bordure(fin=True)

        # Col C – Total dépenses
        c = ws.cell(ligne, 3, f"={ref_dep}")
        c.number_format = '#,##0.00 €'
        c.font      = police(couleur=COULEURS["depense"])
        c.fill      = fill(bg)
        c.alignment = centrer()
        c.border    = style_bordure(fin=True)

        # Col D – Solde annuel
        c = ws.cell(ligne, 4, f"=B{ligne}-C{ligne}")
        c.number_format = '#,##0.00 €'
        c.font      = police(gras=True)
        c.fill      = fill(bg)
        c.alignment = centrer()
        c.border    = style_bordure(fin=True)

        # Col E – Solde cumulé
        if ligne == ligne_debut_donnees:
            formule_cumul = f"=D{ligne}"
        else:
            formule_cumul = f"=E{ligne-1}+D{ligne}"
        c = ws.cell(ligne, 5, formule_cumul)
        c.number_format = '#,##0.00 €'
        c.font      = police(gras=True)
        c.fill      = fill(bg)
        c.alignment = centrer()
        c.border    = style_bordure(fin=True)

        # Col F – Évolution (vs année précédente)
        if ligne == ligne_debut_donnees:
            c = ws.cell(ligne, 6, "—")
            c.alignment = centrer()
            c.fill = fill(bg)
            c.border = style_bordure(fin=True)
        else:
            c = ws.cell(ligne, 6,
                        f'=IFERROR((D{ligne}-D{ligne-1})/ABS(D{ligne-1}),"")')
            c.number_format = '+0.0%;-0.0%;0.0%'
            c.font      = police(gras=True)
            c.fill      = fill(bg)
            c.alignment = centrer()
            c.border    = style_bordure(fin=True)

        lignes_data[annee] = ligne
        ws.row_dimensions[ligne].height = 20
        ligne += 1

    ligne_fin_donnees = ligne - 1

    # ─── LIGNE TOTAUX / MOYENNES ─────────────────────────────────────────────
    ligne += 1
    ws.merge_cells(f"A{ligne}:A{ligne}")
    c = ws.cell(ligne, 1, "TOTAUX")
    c.font      = police(gras=True, couleur="FFFFFF")
    c.fill      = fill(COULEURS["synthese"])
    c.alignment = centrer()
    c.border    = style_bordure()

    for col, formule in [
        (2, f"=SUM(B{ligne_debut_donnees}:B{ligne_fin_donnees})"),
        (3, f"=SUM(C{ligne_debut_donnees}:C{ligne_fin_donnees})"),
        (4, f"=SUM(D{ligne_debut_donnees}:D{ligne_fin_donnees})"),
        (5, f"=E{ligne_fin_donnees}"),  # solde cumulé final
    ]:
        c = ws.cell(ligne, col, formule)
        c.number_format = '#,##0.00 €'
        c.font      = police(gras=True, couleur="FFFFFF")
        c.fill      = fill(COULEURS["synthese"])
        c.alignment = centrer()
        c.border    = style_bordure()

    ws.row_dimensions[ligne].height = 22
    ligne_totaux = ligne
    ligne += 1

    # Moyennes
    ws.merge_cells(f"A{ligne}:A{ligne}")
    c = ws.cell(ligne, 1, "MOYENNES")
    c.font      = police(gras=True, couleur="FFFFFF")
    c.fill      = fill(COULEURS["titre"])
    c.alignment = centrer()
    c.border    = style_bordure()

    for col, formule in [
        (2, f"=AVERAGE(B{ligne_debut_donnees}:B{ligne_fin_donnees})"),
        (3, f"=AVERAGE(C{ligne_debut_donnees}:C{ligne_fin_donnees})"),
        (4, f"=AVERAGE(D{ligne_debut_donnees}:D{ligne_fin_donnees})"),
    ]:
        c = ws.cell(ligne, col, formule)
        c.number_format = '#,##0.00 €'
        c.font      = police(gras=True, couleur="FFFFFF")
        c.fill      = fill(COULEURS["titre"])
        c.alignment = centrer()
        c.border    = style_bordure()

    ws.row_dimensions[ligne].height = 22
    ligne += 2

    # ─── GRAPHIQUE ───────────────────────────────────────────────────────────
    if annees:
        nb = len(annees)

        # BarChart recettes/dépenses
        bar = BarChart()
        bar.type    = "col"
        bar.grouping = "clustered"
        bar.title   = "Recettes & Dépenses par année"
        bar.y_axis.title = "Montant (€)"
        bar.x_axis.title = "Année"
        bar.style   = 10
        bar.width   = 20
        bar.height  = 12

        data_rec = Reference(ws,
                             min_col=2, max_col=2,
                             min_row=ligne_debut_donnees - 1,
                             max_row=ligne_fin_donnees)
        data_dep = Reference(ws,
                             min_col=3, max_col=3,
                             min_row=ligne_debut_donnees - 1,
                             max_row=ligne_fin_donnees)
        cats = Reference(ws,
                         min_col=1,
                         min_row=ligne_debut_donnees,
                         max_row=ligne_fin_donnees)

        bar.add_data(data_rec, titles_from_data=True)
        bar.add_data(data_dep, titles_from_data=True)
        bar.set_categories(cats)
        bar.series[0].graphicalProperties.solidFill = COULEURS["recette"]
        bar.series[1].graphicalProperties.solidFill = COULEURS["depense"]

        ws.add_chart(bar, f"A{ligne}")

        # LineChart solde cumulé
        line = LineChart()
        line.title  = "Évolution du solde cumulé"
        line.y_axis.title = "Solde cumulé (€)"
        line.x_axis.title = "Année"
        line.style  = 10
        line.width  = 20
        line.height = 12

        data_cumul = Reference(ws,
                               min_col=5, max_col=5,
                               min_row=ligne_debut_donnees - 1,
                               max_row=ligne_fin_donnees)
        line.add_data(data_cumul, titles_from_data=True)
        line.set_categories(cats)
        line.series[0].graphicalProperties.line.solidFill = COULEURS["synthese"]
        line.series[0].graphicalProperties.line.width = 25000  # épaisseur

        ws.add_chart(line, f"D{ligne}")


# ============================================================
# GESTION DU FICHIER (CRÉATION OU MISE À JOUR)
# ============================================================
def generer_ou_mettre_a_jour(recettes, depenses, output: Path,
                              annee: int, force: bool):
    NOM_ONGLET = f"Comptes {annee}"
    NOM_SYNTHESE = "📈 Synthèse"

    # ── Ouvrir ou créer le classeur ──────────────────────────────────────────
    if output.exists():
        print(f"📂 Fichier existant détecté : {output.name}")
        wb = openpyxl.load_workbook(output)

        if NOM_ONGLET in wb.sheetnames:
            if not force:
                rep = input(
                    f"⚠️  L'onglet '{NOM_ONGLET}' existe déjà. "
                    f"Le remplacer ? [o/N] "
                ).strip().lower()
                if rep not in ("o", "oui", "y", "yes"):
                    print("❌ Opération annulée.")
                    return
            # Supprimer l'ancien onglet
            del wb[NOM_ONGLET]
            print(f"🗑️  Onglet '{NOM_ONGLET}' supprimé.")

    else:
        wb = openpyxl.Workbook()
        # Supprimer la feuille vide par défaut
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        print(f"🆕 Création du fichier : {output.name}")

    # ── Créer l'onglet annuel ────────────────────────────────────────────────
    ws = wb.create_sheet(NOM_ONGLET)
    ecrire_onglet_annuel(ws, recettes, depenses, annee)
    print(f"✅ Onglet '{NOM_ONGLET}' créé.")

    # ── Trier les onglets annuels par année ──────────────────────────────────
    onglets_annuels = sorted(
        [n for n in wb.sheetnames if n.startswith("Comptes ")],
        key=lambda n: int(n.split()[-1])
    )
    # Remettre dans l'ordre (synthèse en 1er, puis années croissantes)
    ordre = [NOM_SYNTHESE] + onglets_annuels if NOM_SYNTHESE in wb.sheetnames \
        else onglets_annuels
    wb._sheets.sort(key=lambda s: ordre.index(s.title) if s.title in ordre else 99)

    # ── Mettre à jour la synthèse ────────────────────────────────────────────
    try:
        ecrire_onglet_synthese(wb)
        print(f"📈 Onglet '{NOM_SYNTHESE}' mis à jour.")
    except Exception as e:
        print(f"⚠️ Erreur lors de la création de la synthèse: {e}")
        print("Le fichier sera sauvegardé sans la synthèse.")

    try:
        wb.save(output)
        print(f"💾 Fichier sauvegardé : {output}")
    except Exception as e:
        print(f"❌ Erreur lors de la sauvegarde: {e}")
        sys.exit(1)


# ============================================================
# MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser(
        description="Génère un livre de comptes Excel depuis des relevés PDF."
    )
    parser.add_argument("pdfs", nargs="+", help="Fichiers PDF à traiter")
    parser.add_argument("-o", "--output", default="livre_comptes.xlsx",
                        help="Fichier Excel de sortie (défaut: livre_comptes.xlsx)")
    parser.add_argument("-a", "--annee", type=int,
                        default=datetime.now().year,
                        help="Année des relevés (défaut: année en cours)")
    parser.add_argument("-f", "--force", action="store_true",
                        help="Remplacer l'onglet existant sans confirmation")
    args = parser.parse_args()
    args.output = Path(args.output)

    print(f"\n📒 Livre de comptes — Année {args.annee}")
    print("=" * 50)

    # Vérification des PDF
    pdfs_valides = []
    for p in args.pdfs:
        path = Path(p)
        if not path.exists():
            print(f"⚠️  Fichier introuvable : {p}")
        elif path.suffix.lower() != ".pdf":
            print(f"⚠️  Pas un PDF : {p}")
        else:
            pdfs_valides.append(path)

    if not pdfs_valides:
        print("❌ Aucun fichier PDF valide trouvé.")
        sys.exit(1)

    # Extraction
    print(f"\n🔍 Extraction depuis {len(pdfs_valides)} fichier(s)...")
    all_recettes, all_depenses = [], []
    for pdf in sorted(pdfs_valides):
        print(f"  📄 {pdf.name}...")
        try:
            rec, dep = extraire_operations_pdf_camelot(pdf, args.annee)
            print(f"     → {len(rec)} recette(s), {len(dep)} dépense(s)")
            all_recettes.extend(rec)
            all_depenses.extend(dep)
        except Exception as e:
            print(f"     ❌ Erreur lors de l'extraction du PDF {pdf.name}: {e}")
            continue

    def cle_tri(op):
        try:
            return datetime.strptime(op[0], "%d/%m/%Y")
        except ValueError:
            return datetime.min

    all_recettes.sort(key=cle_tri)
    all_depenses.sort(key=cle_tri)

    print(f"\n📊 Total : {len(all_recettes)} recettes, {len(all_depenses)} dépenses")

    generer_ou_mettre_a_jour(all_recettes, all_depenses,
                             args.output, args.annee, args.force)


if __name__ == "__main__":
    main()
