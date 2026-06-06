"""Excel writer module for generating bookkeeping reports."""
import logging
from pathlib import Path
from datetime import datetime
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

from models import Transaction
from config import COLORS, INCOME_CATEGORIES, EXPENSE_CATEGORIES

logger = logging.getLogger(__name__)


def style_border(thin: bool = False):
    """Create a border style."""
    style = "thin" if thin else "hair"
    return Border(left=Side(style=style), right=Side(style=style), top=Side(style=style), bottom=Side(style=style))


def fill(color: str):
    """Create a pattern fill."""
    return PatternFill("solid", fgColor=color)


def font(bold: bool = False, size: int = 11, color: str = "000000", italic: bool = False):
    """Create a font style."""
    return Font(bold=bold, size=size, color=color, italic=italic)


def center(horizontal: str = "center", vertical: str = "center", wrap: bool = False):
    """Create an alignment."""
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


def write_annual_sheet(ws, income_transactions: List[Transaction], expense_transactions: List[Transaction], year: int):
    """Fill a sheet with incomes, expenses, summary and breakdown."""
    
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    widths = {"A": 14, "B": 14, "C": 45, "D": 18, "E": 14, "F": 22, "G": 20}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    headers = ["Date opération", "Date valeur", "Libellé", "Référence",
               "Montant (€)", "Catégorie", "Source"]

    row = 1

    # General title
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row, 1, f"📒  LIVRE DE COMPTES {year}")
    c.font = font(bold=True, size=16, color="FFFFFF")
    c.fill = fill(COLORS["titre"])
    c.alignment = center()
    c.border = style_border()
    ws.row_dimensions[row].height = 32
    row += 2

    # INCOMES section
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row, 1, "💰  RECETTES")
    c.font = font(bold=True, size=13, color="FFFFFF")
    c.fill = fill(COLORS["recette"])
    c.alignment = center()
    c.border = style_border()
    ws.row_dimensions[row].height = 26
    row += 1

    # Income headers
    header_income_row = row
    for col, title in enumerate(headers, 1):
        c = ws.cell(row, col, title)
        c.font = font(bold=True, color="FFFFFF")
        c.fill = fill(COLORS["recette"])
        c.alignment = center()
        c.border = style_border()
    ws.row_dimensions[row].height = 20
    row += 1

    # Income data
    start_income_row = row
    cats_income_dv = ",".join(INCOME_CATEGORIES)
    dv_income = DataValidation(type="list", formula1=f'"{cats_income_dv}"', showDropDown=False)
    ws.add_data_validation(dv_income)

    for i, op in enumerate(income_transactions):
        bg = COLORS["blanc"] if i % 2 == 0 else COLORS["alternance"]
        values = [
            op.date_operation,
            op.date_valeur,
            op.libelle,
            op.reference,
            op.montant,
            op.categorie,
            op.source
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row, col, val)
            c.fill = fill(bg)
            c.border = style_border(thin=True)
            c.alignment = center("left") if col == 3 else center()
            if col == 5:
                c.number_format = '#,##0.00 €'
                c.font = font(color=COLORS["recette"])
        dv_income.add(ws.cell(row, 6))
        row += 1

    end_income_row = row - 1

    # Total income
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row, 1, "TOTAL RECETTES")
    c.font = font(bold=True, color="FFFFFF")
    c.fill = fill(COLORS["recette"])
    c.alignment = center("right")
    c.border = style_border()

    col_total_income = 5
    cell_total_income = ws.cell(row, col_total_income)
    if income_transactions:
        cell_total_income.value = f"=SUM(E{start_income_row}:E{end_income_row})"
    else:
        cell_total_income.value = 0
    cell_total_income.font = font(bold=True, color="FFFFFF")
    cell_total_income.fill = fill(COLORS["recette"])
    cell_total_income.number_format = '#,##0.00 €'
    cell_total_income.alignment = center()
    cell_total_income.border = style_border()
    ref_total_income = f"E{row}"
    ws.row_dimensions[row].height = 22
    row += 2

    # EXPENSES section
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row, 1, "💸  DÉPENSES")
    c.font = font(bold=True, size=13, color="FFFFFF")
    c.fill = fill(COLORS["depense"])
    c.alignment = center()
    c.border = style_border()
    ws.row_dimensions[row].height = 26
    row += 1

    # Expense headers
    for col, title in enumerate(headers, 1):
        c = ws.cell(row, col, title)
        c.font = font(bold=True, color="FFFFFF")
        c.fill = fill(COLORS["depense"])
        c.alignment = center()
        c.border = style_border()
    ws.row_dimensions[row].height = 20
    row += 1

    # Expense data
    start_expense_row = row
    cats_expense_dv = ",".join(EXPENSE_CATEGORIES)
    dv_expense = DataValidation(type="list", formula1=f'"{cats_expense_dv}"', showDropDown=False)
    ws.add_data_validation(dv_expense)

    for i, op in enumerate(expense_transactions):
        bg = COLORS["blanc"] if i % 2 == 0 else COLORS["alternance"]
        values = [
            op.date_operation,
            op.date_valeur,
            op.libelle,
            op.reference,
            op.montant,
            op.categorie,
            op.source
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row, col, val)
            c.fill = fill(bg)
            c.border = style_border(thin=True)
            c.alignment = center("left") if col == 3 else center()
            if col == 5:
                c.number_format = '#,##0.00 €'
                c.font = font(color=COLORS["depense"])
        dv_expense.add(ws.cell(row, 6))
        row += 1

    end_expense_row = row - 1

    # Total expenses
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row, 1, "TOTAL DÉPENSES")
    c.font = font(bold=True, color="FFFFFF")
    c.fill = fill(COLORS["depense"])
    c.alignment = center("right")
    c.border = style_border()

    cell_total_expense = ws.cell(row, 5)
    if expense_transactions:
        cell_total_expense.value = f"=SUM(E{start_expense_row}:E{end_expense_row})"
    else:
        cell_total_expense.value = 0
    cell_total_expense.font = font(bold=True, color="FFFFFF")
    cell_total_expense.fill = fill(COLORS["depense"])
    cell_total_expense.number_format = '#,##0.00 €'
    cell_total_expense.alignment = center()
    cell_total_expense.border = style_border()
    ref_total_expense = f"E{row}"
    ws.row_dimensions[row].height = 22
    row += 2

    # Annual summary
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row, 1, "📊  RÉCAPITULATIF ANNUEL")
    c.font = font(bold=True, size=13, color="FFFFFF")
    c.fill = fill(COLORS["recap"])
    c.alignment = center()
    c.border = style_border()
    ws.row_dimensions[row].height = 26
    row += 1

    summary_items = [
        ("Total recettes", f"={ref_total_income}", COLORS["recette_clair"], COLORS["recette"]),
        ("Total dépenses", f"={ref_total_expense}", COLORS["depense_clair"], COLORS["depense"]),
        ("Solde de l'année", f"={ref_total_income}-{ref_total_expense}", COLORS["recap_clair"], COLORS["recap"]),
    ]

    for label, formula, bg_light, bg_dark in summary_items:
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row, 1, label)
        c.font = font(bold=True, color="FFFFFF")
        c.fill = fill(bg_dark)
        c.alignment = center("right")
        c.border = style_border()

        c2 = ws.cell(row, 5, formula)
        c2.font = font(bold=True)
        c2.fill = fill(bg_light)
        c2.number_format = '#,##0.00 €'
        c2.alignment = center()
        c2.border = style_border()
        ws.row_dimensions[row].height = 22
        row += 1

    row += 1

    # Breakdown by category
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row, 1, "🎯  VENTILATION PAR CATÉGORIE")
    c.font = font(bold=True, size=13, color="FFFFFF")
    c.fill = fill(COLORS["recap"])
    c.alignment = center()
    c.border = style_border()
    ws.row_dimensions[row].height = 26
    row += 1

    # Income by category
    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row, 1, "Catégorie recette")
    c.font = font(bold=True, color="FFFFFF")
    c.fill = fill(COLORS["recette"])
    c.alignment = center()
    c.border = style_border()

    ws.merge_cells(f"D{row}:E{row}")
    c = ws.cell(row, 4, "Montant (€)")
    c.font = font(bold=True, color="FFFFFF")
    c.fill = fill(COLORS["recette"])
    c.alignment = center()
    c.border = style_border()
    row += 1

    for cat in INCOME_CATEGORIES:
        if income_transactions:
            formula = (f'=SUMIF(F{start_income_row}:F{end_income_row},'
                       f'"{cat}",E{start_income_row}:E{end_income_row})')
        else:
            formula = 0
        ws.merge_cells(f"A{row}:C{row}")
        c = ws.cell(row, 1, cat)
        c.fill = fill(COLORS["recette_clair"])
        c.border = style_border(thin=True)
        c.alignment = center("left")

        ws.merge_cells(f"D{row}:E{row}")
        c2 = ws.cell(row, 4, formula)
        c2.number_format = '#,##0.00 €'
        c2.fill = fill(COLORS["recette_clair"])
        c2.border = style_border(thin=True)
        c2.alignment = center()
        row += 1

    row += 1

    # Expenses by category
    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row, 1, "Catégorie dépense")
    c.font = font(bold=True, color="FFFFFF")
    c.fill = fill(COLORS["depense"])
    c.alignment = center()
    c.border = style_border()

    ws.merge_cells(f"D{row}:E{row}")
    c = ws.cell(row, 4, "Montant (€)")
    c.font = font(bold=True, color="FFFFFF")
    c.fill = fill(COLORS["depense"])
    c.alignment = center()
    c.border = style_border()
    row += 1

    for cat in EXPENSE_CATEGORIES:
        if expense_transactions:
            formula = (f'=SUMIF(F{start_expense_row}:F{end_expense_row},'
                       f'"{cat}",E{start_expense_row}:E{end_expense_row})')
        else:
            formula = 0
        ws.merge_cells(f"A{row}:C{row}")
        c = ws.cell(row, 1, cat)
        c.fill = fill(COLORS["depense_clair"])
        c.border = style_border(thin=True)
        c.alignment = center("left")

        ws.merge_cells(f"D{row}:E{row}")
        c2 = ws.cell(row, 4, formula)
        c2.number_format = '#,##0.00 €'
        c2.fill = fill(COLORS["depense_clair"])
        c2.border = style_border(thin=True)
        c2.alignment = center()
        row += 1


def write_summary_sheet(wb: openpyxl.Workbook):
    """Recreate the '📈 Synthèse' sheet by reading totals from each annual sheet."""
    
    SUMMARY_SHEET_NAME = "📈 Synthèse"
    if SUMMARY_SHEET_NAME in wb.sheetnames:
        del wb[SUMMARY_SHEET_NAME]

    ws = wb.create_sheet(SUMMARY_SHEET_NAME, 0)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    annual_sheets = sorted(
        [n for n in wb.sheetnames if n.startswith("Comptes ")],
        key=lambda n: int(n.split()[-1])
    )
    years = [int(n.split()[-1]) for n in annual_sheets]

    def find_summary_cells(ws_annual):
        """Return (income_total_cell, expense_total_cell) by searching for labels."""
        income_cell = expense_cell = None
        for row in ws_annual.iter_rows():
            for cell in row:
                if cell.value == "TOTAL RECETTES":
                    income_cell = ws_annual.cell(cell.row, 5)
                if cell.value == "TOTAL DÉPENSES":
                    expense_cell = ws_annual.cell(cell.row, 5)
        return income_cell, expense_cell

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 16

    row = 1

    # Title
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row, 1, "📈  SYNTHÈSE PLURIANNUELLE")
    c.font = font(bold=True, size=16, color="FFFFFF")
    c.fill = fill(COLORS["synthese"])
    c.alignment = center()
    c.border = style_border()
    ws.row_dimensions[row].height = 34
    row += 1

    # Subtitle with update date
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row, 1, f"Mis à jour le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
    c.font = font(italic=True, size=10, color="FFFFFF")
    c.fill = fill(COLORS["synthese"])
    c.alignment = center()
    ws.row_dimensions[row].height = 16
    row += 1

    # Headers
    headers_syn = ["Année", "Total recettes", "Total dépenses", "Solde annuel", "Solde cumulé", "Évolution"]
    colors_headers = [
        COLORS["synthese"], COLORS["recette"], COLORS["depense"],
        COLORS["recap"], COLORS["recap"], COLORS["titre"],
    ]
    for col, (title, color) in enumerate(zip(headers_syn, colors_headers), 1):
        c = ws.cell(row, col, title)
        c.font = font(bold=True, color="FFFFFF")
        c.fill = fill(color)
        c.alignment = center()
        c.border = style_border()
    ws.row_dimensions[row].height = 22
    row += 1

    # Data by year
    start_data_row = row

    for idx, (year, sheet_name) in enumerate(zip(years, annual_sheets)):
        ws_ann = wb[sheet_name]
        income_cell, expense_cell = find_summary_cells(ws_ann)

        name_safe = sheet_name.replace("'", "''")
        if income_cell and expense_cell:
            ref_income = f"'{name_safe}'!E{income_cell.row}"
            ref_expense = f"'{name_safe}'!E{expense_cell.row}"
        else:
            ref_income = ref_expense = "0"

        bg = COLORS["blanc"] if idx % 2 == 0 else COLORS["alternance"]

        # Year
        c = ws.cell(row, 1, year)
        c.font = font(bold=True)
        c.fill = fill(COLORS["synthese_clair"])
        c.alignment = center()
        c.border = style_border()

        # Total income
        c = ws.cell(row, 2, f"={ref_income}")
        c.number_format = '#,##0.00 €'
        c.font = font(color=COLORS["recette"])
        c.fill = fill(bg)
        c.alignment = center()
        c.border = style_border(thin=True)

        # Total expenses
        c = ws.cell(row, 3, f"={ref_expense}")
        c.number_format = '#,##0.00 €'
        c.font = font(color=COLORS["depense"])
        c.fill = fill(bg)
        c.alignment = center()
        c.border = style_border(thin=True)

        # Annual balance
        c = ws.cell(row, 4, f"=B{row}-C{row}")
        c.number_format = '#,##0.00 €'
        c.font = font(bold=True)
        c.fill = fill(bg)
        c.alignment = center()
        c.border = style_border(thin=True)

        # Cumulative balance
        if row == start_data_row:
            formula_cumul = f"=D{row}"
        else:
            formula_cumul = f"=E{row-1}+D{row}"
        c = ws.cell(row, 5, formula_cumul)
        c.number_format = '#,##0.00 €'
        c.font = font(bold=True)
        c.fill = fill(bg)
        c.alignment = center()
        c.border = style_border(thin=True)

        # Evolution (vs previous year)
        if row == start_data_row:
            c = ws.cell(row, 6, "—")
            c.alignment = center()
            c.fill = fill(bg)
            c.border = style_border(thin=True)
        else:
            c = ws.cell(row, 6, f'=IFERROR((D{row}-D{row-1})/ABS(D{row-1}),"")') 
            c.number_format = '+0.0%;-0.0%;0.0%'
            c.font = font(bold=True)
            c.fill = fill(bg)
            c.alignment = center()
            c.border = style_border(thin=True)

        ws.row_dimensions[row].height = 20
        row += 1

    end_data_row = row - 1

    # Totals / averages row
    row += 1
    ws.merge_cells(f"A{row}:A{row}")
    c = ws.cell(row, 1, "TOTAUX")
    c.font = font(bold=True, color="FFFFFF")
    c.fill = fill(COLORS["synthese"])
    c.alignment = center()
    c.border = style_border()

    for col, formula in [
        (2, f"=SUM(B{start_data_row}:B{end_data_row})"),
        (3, f"=SUM(C{start_data_row}:C{end_data_row})"),
        (4, f"=SUM(D{start_data_row}:D{end_data_row})"),
        (5, f"=E{end_data_row}"),
    ]:
        c = ws.cell(row, col, formula)
        c.number_format = '#,##0.00 €'
        c.font = font(bold=True, color="FFFFFF")
        c.fill = fill(COLORS["synthese"])
        c.alignment = center()
        c.border = style_border()

    ws.row_dimensions[row].height = 22
    row += 1

    # Averages
    ws.merge_cells(f"A{row}:A{row}")
    c = ws.cell(row, 1, "MOYENNES")
    c.font = font(bold=True, color="FFFFFF")
    c.fill = fill(COLORS["titre"])
    c.alignment = center()
    c.border = style_border()

    for col, formula in [
        (2, f"=AVERAGE(B{start_data_row}:B{end_data_row})"),
        (3, f"=AVERAGE(C{start_data_row}:C{end_data_row})"),
        (4, f"=AVERAGE(D{start_data_row}:D{end_data_row})"),
    ]:
        c = ws.cell(row, col, formula)
        c.number_format = '#,##0.00 €'
        c.font = font(bold=True, color="FFFFFF")
        c.fill = fill(COLORS["titre"])
        c.alignment = center()
        c.border = style_border()

    ws.row_dimensions[row].height = 22
    row += 2

    # Chart
    if years:
        bar = BarChart()
        bar.type = "col"
        bar.grouping = "clustered"
        bar.title = "Recettes & Dépenses par année"
        bar.y_axis.title = "Montant (€)"
        bar.x_axis.title = "Année"
        bar.style = 10
        bar.width = 20
        bar.height = 12

        data_income = Reference(ws, min_col=2, max_col=2, 
                           min_row=start_data_row - 1, max_row=end_data_row)
        data_expense = Reference(ws, min_col=3, max_col=3,
                           min_row=start_data_row - 1, max_row=end_data_row)
        categories = Reference(ws, min_col=1,
                        min_row=start_data_row, max_row=end_data_row)

        bar.add_data(data_income, titles_from_data=True)
        bar.add_data(data_expense, titles_from_data=True)
        bar.set_categories(categories)
        bar.series[0].graphicalProperties.solidFill = COLORS["recette"]
        bar.series[1].graphicalProperties.solidFill = COLORS["depense"]

        ws.add_chart(bar, f"A{row}")

        # LineChart cumulative balance
        line = LineChart()
        line.title = "Évolution du solde cumulé"
        line.y_axis.title = "Solde cumulé (€)"
        line.x_axis.title = "Année"
        line.style = 10
        line.width = 20
        line.height = 12

        data_cumul = Reference(ws, min_col=5, max_col=5,
                             min_row=start_data_row - 1, max_row=end_data_row)
        line.add_data(data_cumul, titles_from_data=True)
        line.set_categories(categories)
        line.series[0].graphicalProperties.line.solidFill = COLORS["synthese"]
        line.series[0].graphicalProperties.line.width = 25000

        ws.add_chart(line, f"D{row}")


def generate_or_update_excel(income_transactions: List[Transaction], 
                           expense_transactions: List[Transaction], 
                           output_path: Path, 
                           year: int, 
                           force: bool):
    """Generate or update the Excel file with the transactions."""
    SHEET_NAME = f"Comptes {year}"
    SUMMARY_SHEET_NAME = "📈 Synthèse"

    if output_path.exists():
        logger.info(f"Existing file detected: {output_path.name}")
        wb = openpyxl.load_workbook(output_path)

        if SHEET_NAME in wb.sheetnames:
            if not force:
                # Since we're not in interactive mode, we'll proceed with update
                logger.info(f"Sheet '{SHEET_NAME}' exists, updating it.")
            del wb[SHEET_NAME]
            logger.info(f"Deleted sheet '{SHEET_NAME}'.")
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        logger.info(f"Creating new file: {output_path.name}")

    ws = wb.create_sheet(SHEET_NAME)
    write_annual_sheet(ws, income_transactions, expense_transactions, year)
    logger.info(f"Created sheet '{SHEET_NAME}'.")

    annual_sheets = sorted([n for n in wb.sheetnames if n.startswith("Comptes ")],
                          key=lambda n: int(n.split()[-1]))
    order = [SUMMARY_SHEET_NAME] + annual_sheets if SUMMARY_SHEET_NAME in wb.sheetnames else annual_sheets
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)

    try:
        write_summary_sheet(wb)
        logger.info(f"Updated sheet '{SUMMARY_SHEET_NAME}'.")
    except Exception as e:
        logger.error(f"Error creating summary sheet: {e}")
        logger.info("File will be saved without summary sheet.")

    try:
        wb.save(output_path)
        logger.info(f"Saved file: {output_path}")
    except Exception as e:
        logger.error(f"Error saving file: {e}")
        raise