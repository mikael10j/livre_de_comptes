#!/usr/bin/env python3
"""
Copyright 2026 mikael10j

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

    http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
"""

"""
Tests d'intégration avec les vrais fichiers PDF.
Ces tests utilisent les fichiers PDF réels présents dans le répertoire parent.
"""

import unittest
import tempfile
import shutil
import sys
from pathlib import Path
import openpyxl

# Ajouter le répertoire parent au path pour l'import
sys.path.append(str(Path(__file__).parent.parent))
from parser import extract_operations_pdf_camelot
from excel_writer import generate_or_update_excel
from config import INCOME_CATEGORIES, EXPENSE_CATEGORIES
from models import Transaction


class TestIntegrationReal(unittest.TestCase):
    """Tests d'intégration avec les vrais fichiers PDF."""

    def setUp(self):
        """Configuration des tests d'intégration."""
        self.temp_dir = tempfile.mkdtemp()
        self.temp_path = Path(self.temp_dir)
        
        # Trouver les fichiers PDF réels dans le répertoire parent
        parent_dir = Path(__file__).parent.parent
        self.pdf_files = list(parent_dir.glob("*.pdf"))
        
    def tearDown(self):
        """Nettoyage après tests."""
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_extraction_tous_pdfs_reels(self):
        """Test d'extraction avec tous les PDFs réels disponibles."""
        if not self.pdf_files:
            self.skipTest("Aucun fichier PDF trouvé pour les tests")
        
        total_operations = 0
        total_recettes = 0
        total_depenses = 0
        
        for pdf_file in self.pdf_files:
            with self.subTest(pdf=pdf_file.name):
                # Déterminer l'année selon le nom du fichier
                annee = 2023 if "2023" in pdf_file.name else 2019
                
                result = extract_operations_pdf_camelot(pdf_file, annee)
                
                # Vérifications de base
                self.assertIsNotNone(result)
                self.assertIsInstance(result.income_transactions, list)
                self.assertIsInstance(result.expense_transactions, list)
                
                # Chaque opération doit avoir la bonne structure
                for recette in result.income_transactions:
                    self.assertIsInstance(recette, Transaction)
                    self.assertIsInstance(recette.montant, (int, float))  # Montant numérique
                    self.assertGreater(recette.montant, 0)  # Montant positif
                
                for depense in result.expense_transactions:
                    self.assertIsInstance(depense, Transaction)
                    self.assertIsInstance(depense.montant, (int, float))
                    self.assertGreater(depense.montant, 0)
                
                total_operations += len(result.income_transactions) + len(result.expense_transactions)
                total_recettes += len(result.income_transactions)
                total_depenses += len(result.expense_transactions)
        
        print(f"\n📊 Extraction réelle: {total_operations} opérations "
              f"({total_recettes} recettes, {total_depenses} dépenses)")
        
        # Vérifier qu'au moins quelques opérations ont été extraites
        self.assertGreater(total_operations, 0, "Aucune opération extraite des PDFs réels")

    def test_generation_excel_complete(self):
        """Test de génération complète d'un fichier Excel avec tous les PDFs."""
        if not self.pdf_files:
            self.skipTest("Aucun fichier PDF trouvé pour les tests")
        
        output_path = self.temp_path / "test_integration_complete.xlsx"
        
        # Traiter les PDFs par année
        annees_traitees = set()
        
        for pdf_file in self.pdf_files:
            annee = 2023 if "2023" in pdf_file.name else 2019
            
            if annee in annees_traitees:
                continue  # Éviter de traiter plusieurs fois la même année
                
            result = extract_operations_pdf_camelot(pdf_file, annee)
            
            if result.income_transactions or result.expense_transactions:  # Seulement si des opérations ont été trouvées
                generate_or_update_excel(result.income_transactions, result.expense_transactions, output_path, annee, force=True)
                annees_traitees.add(annee)
        
        # Vérifier le fichier généré
        self.assertTrue(output_path.exists())
        
        # Vérifier le contenu Excel
        wb = openpyxl.load_workbook(output_path)
        
        # Doit contenir au moins un onglet annuel
        onglets_comptes = [sheet for sheet in wb.sheetnames if sheet.startswith("Comptes ")]
        self.assertGreater(len(onglets_comptes), 0, "Aucun onglet de comptes créé")
        
        # Doit contenir l'onglet synthèse
        self.assertIn("📈 Synthèse", wb.sheetnames)
        
        print(f"\n📊 Excel généré avec {len(onglets_comptes)} onglet(s) : {onglets_comptes}")

    def test_coherence_donnees_extraites(self):
        """Test de cohérence des données extraites."""
        if not self.pdf_files:
            self.skipTest("Aucun fichier PDF trouvé pour les tests")
        
        for pdf_file in self.pdf_files:
            with self.subTest(pdf=pdf_file.name):
                annee = 2023 if "2023" in pdf_file.name else 2019
                result = extract_operations_pdf_camelot(pdf_file, annee)
                
                # Vérifier la cohérence des dates
                for operation in result.income_transactions + result.expense_transactions:
                    self.assertIsInstance(operation, Transaction)
                    
                    # Les dates doivent contenir l'année
                    self.assertIn(str(annee), operation.date_operation)
                    self.assertIn(str(annee), operation.date_valeur)
                    
                    # Le libellé ne doit pas être vide
                    self.assertGreater(len(operation.libelle.strip()), 0)
                    
                    # Le montant doit être positif
                    self.assertGreater(operation.montant, 0)
                    
                    # La catégorie doit être valide
                    if operation.is_income:
                        self.assertIn(operation.categorie, INCOME_CATEGORIES)
                    else:
                        self.assertIn(operation.categorie, EXPENSE_CATEGORIES)
                    
                    # La source doit correspondre au fichier
                    self.assertEqual(operation.source, pdf_file.stem)

    def test_reproductibilite(self):
        """Test de reproductibilité : même PDF doit donner mêmes résultats."""
        if not self.pdf_files:
            self.skipTest("Aucun fichier PDF trouvé pour les tests")
        
        pdf_test = self.pdf_files[0]  # Prendre le premier PDF
        annee = 2023 if "2023" in pdf_test.name else 2019
        
        # Extraire deux fois
        result1 = extract_operations_pdf_camelot(pdf_test, annee)
        result2 = extract_operations_pdf_camelot(pdf_test, annee)
        
        # Les résultats doivent être identiques
        self.assertEqual(len(result1.income_transactions), len(result2.income_transactions))
        self.assertEqual(len(result1.expense_transactions), len(result2.expense_transactions))
        
        # Comparer operation par operation
        for op1, op2 in zip(result1.income_transactions, result2.income_transactions):
            self.assertEqual(op1, op2)
            
        for op1, op2 in zip(result1.expense_transactions, result2.expense_transactions):
            self.assertEqual(op1, op2)
        
        print(f"✅ Reproductibilité confirmée pour {pdf_test.name}")

    def test_gestion_erreurs_fichiers(self):
        """Test de gestion d'erreurs avec des fichiers problématiques."""
        # Tester avec un fichier inexistant
        pdf_inexistant = Path("fichier_inexistant.pdf")
        result = extract_operations_pdf_camelot(pdf_inexistant, 2023)
        
        # Doit retourner des listes vides sans lever d'exception
        self.assertEqual(result.income_transactions, [])
        self.assertEqual(result.expense_transactions, [])
        
        # Tester avec un fichier vide
        fichier_vide = self.temp_path / "vide.pdf"
        fichier_vide.touch()
        
        result = extract_operations_pdf_camelot(fichier_vide, 2023)
        self.assertEqual(result.income_transactions, [])
        self.assertEqual(result.expense_transactions, [])


def run_integration_tests():
    """Exécute les tests d'intégration avec les vrais fichiers."""
    print("🔗 TESTS D'INTÉGRATION")
    print("=" * 50)
    
    # Vérifier la présence de fichiers PDF
    parent_dir = Path(__file__).parent.parent
    pdf_files = list(parent_dir.glob("*.pdf"))
    if not pdf_files:
        print("⚠️  Aucun fichier PDF trouvé. Les tests d'intégration seront ignorés.")
        return True
    
    print(f"📄 {len(pdf_files)} fichier(s) PDF trouvé(s): {[f.name for f in pdf_files]}")
    
    suite = unittest.TestLoader().loadTestsFromTestCase(TestIntegrationReal)
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)
    
    print(f"\n📊 Résultats: {result.testsRun - len(result.failures) - len(result.errors)}/{result.testsRun} tests réussis")
    
    return result.wasSuccessful()


if __name__ == "__main__":
    success = run_integration_tests()
    exit(0 if success else 1)