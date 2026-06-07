#!/usr/bin/env python3
"""
Copyright 2026 mikael10j

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

    http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
"""

"""
Tests unitaires pour le script livre_comptes.py

Usage:
    python -m pytest tests/test_unitaires.py -v
    python tests/test_unitaires.py  # Exécution directe
"""

import unittest
import tempfile
import shutil
import sys
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock
import pandas as pd
import openpyxl
from datetime import datetime

# Ajouter le répertoire parent au path pour l'import
sys.path.append(str(Path(__file__).parent.parent))
from parser import parse_amount, find_header_row, create_named_dataframe, extract_operations_pdf_camelot, categorize_transaction
from excel_writer import generate_or_update_excel, write_annual_sheet, style_border, fill, font, center
from config import INCOME_RULES, EXPENSE_RULES, INCOME_CATEGORIES, EXPENSE_CATEGORIES
from models import Transaction


class TestFonctionsUtilitaires(unittest.TestCase):
    """Tests pour les fonctions utilitaires."""

    def test_parse_montant_formats_valides(self):
        """Test du parsing de différents formats de montants."""
        test_cases = [
            ("123,45", 123.45),
            ("1 234,56", 1234.56),
            ("0,00", 0.0),
            ("999999,99", 999999.99),
            ("10", 10.0),
            ("10.50", 10.5),
            ("10,50", 10.5),
        ]
        
        for input_str, expected in test_cases:
            with self.subTest(input_str=input_str):
                result = parse_amount(input_str)
                self.assertEqual(result, expected)

    def test_parse_montant_formats_invalides(self):
        """Test du parsing avec des formats invalides."""
        test_cases = ["", "abc", "12.34.56", None, "   "]
        
        for input_str in test_cases:
            with self.subTest(input_str=input_str):
                result = parse_amount(input_str)
                self.assertEqual(result, 0.0)
                
    def test_parse_montant_nan(self):
        """Test spécifique pour 'nan'."""
        import math
        result = parse_amount("nan")
        # nan peut retourner soit 0.0 soit nan selon l'implémentation
        self.assertTrue(result == 0.0 or math.isnan(result))

    def test_categoriser_recettes(self):
        """Test de la catégorisation automatique des recettes."""
        test_cases = [
            ("Cotisation adhésion 2023", "Cotisations"),
            ("Rem Chq 123456", "Buvette/Tournoi"),
            ("Versement tournoi", "Buvette/Tournoi"),
            ("Don sponsor ABC", "Sponsoring/Dons"),
            ("Subvention mairie", "Subventions"),
            ("VIR RECU sponsor", "Sponsoring/Dons"),
            ("Opération inconnue", "À préciser"),
        ]
        
        for libelle, expected_cat in test_cases:
            with self.subTest(libelle=libelle):
                result = categorize_transaction(libelle, INCOME_RULES)
                self.assertEqual(result, expected_cat)

    def test_categoriser_depenses(self):
        """Test de la catégorisation automatique des dépenses."""
        test_cases = [
            ("Retrait GAB", "Buvette/Tournoi"),
            ("CHQ 123456", "Autres"),
            ("Assurance SMACL", "Assurance"),
            ("Frais bancaires", "Frais bancaires"),
            ("Décathlon matériel", "Matériel sportif"),
            ("Location salle", "Location salle"),
            ("Essence déplacement", "Déplacements"),
            ("Achat buvette", "Buvette/Tournoi"),
            ("Opération inconnue", "À préciser"),
        ]
        
        for libelle, expected_cat in test_cases:
            with self.subTest(libelle=libelle):
                result = categorize_transaction(libelle, EXPENSE_RULES)
                self.assertEqual(result, expected_cat)


class TestExtractionPDF(unittest.TestCase):
    """Tests pour l'extraction des données PDF."""

    def setUp(self):
        """Configuration des tests."""
        # Créer un DataFrame de test simulant une table extraite
        self.df_test = pd.DataFrame([
            ["Titre", "Sous-titre", "", "", "", ""],
            ["", "", "", "", "", ""],
            ["Date", "Date", "Libellé des opérations", "", "Débit", "Crédit"],
            ["opé.", "valeur", "", "", "", ""],
            ["26.01", "26.01", "Chèque 123456", "", "124,65", ""],
            ["27.01", "27.01", "Versement espèces", "", "", "280,00"],
            ["", "", "Total", "124,65", "280,00", ""],
        ])

    def test_find_header_row(self):
        """Test de détection de la ligne d'en-têtes."""
        result = find_header_row(self.df_test)
        self.assertEqual(result, 2)  # Ligne avec "Libellé", "Débit", "Crédit"

    def test_find_header_row_absent(self):
        """Test quand aucune ligne d'en-têtes n'est trouvée."""
        df_sans_entetes = pd.DataFrame([
            ["Données", "quelconques", "sans"],
            ["en-têtes", "valides", "ici"],
        ])
        result = find_header_row(df_sans_entetes)
        self.assertIsNone(result)

    def test_create_named_dataframe(self):
        """Test de création d'un DataFrame avec colonnes nommées."""
        header_row_idx = 2
        result = create_named_dataframe(self.df_test, header_row_idx)
        
        self.assertIsNotNone(result)
        self.assertIn("LIBELLE", result.columns)
        self.assertIn("DEBIT", result.columns)
        self.assertIn("CREDIT", result.columns)
        
        # Vérifier que les données commencent après les en-têtes
        self.assertEqual(len(result), 4)  # 7 lignes totales - 3 lignes d'en-têtes

    def test_create_named_dataframe_header_none(self):
        """Test quand header_row_idx est None."""
        result = create_named_dataframe(self.df_test, None)
        self.assertIsNone(result)


class TestGenerationExcel(unittest.TestCase):
    """Tests pour la génération de fichiers Excel."""

    def setUp(self):
        """Configuration des tests."""
        self.temp_dir = tempfile.mkdtemp()
        self.temp_path = Path(self.temp_dir)
        
        # Données de test
        self.recettes_test = [
            Transaction("01/01/2023", "01/01/2023", "Cotisation", "REF001", 100.0, "Cotisations", "test.pdf", True),
            Transaction("02/01/2023", "02/01/2023", "Tournoi", "REF002", 50.0, "Buvette/Tournoi", "test.pdf", True),
        ]
        self.depenses_test = [
            Transaction("03/01/2023", "03/01/2023", "Matériel", "REF003", 75.0, "Matériel sportif", "test.pdf", False),
            Transaction("04/01/2023", "04/01/2023", "Arbitre", "REF004", 25.0, "Déplacements", "test.pdf", False),
        ]

    def tearDown(self):
        """Nettoyage après tests."""
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_generer_nouveau_fichier(self):
        """Test de génération d'un nouveau fichier Excel."""
        output_path = self.temp_path / "test_nouveau.xlsx"
        
        generate_or_update_excel(
            self.recettes_test, self.depenses_test, 
            output_path, 2023, force=True
        )
        
        # Vérifier que le fichier a été créé
        self.assertTrue(output_path.exists())
        
        # Vérifier le contenu
        wb = openpyxl.load_workbook(output_path)
        self.assertIn("Comptes 2023", wb.sheetnames)
        self.assertIn("📈 Synthèse", wb.sheetnames)

    def test_mise_a_jour_fichier_existant(self):
        """Test de mise à jour d'un fichier existant."""
        output_path = self.temp_path / "test_existant.xlsx"
        
        # Créer un fichier initial
        wb = openpyxl.Workbook()
        wb.save(output_path)
        
        generate_or_update_excel(
            self.recettes_test, self.depenses_test, 
            output_path, 2023, force=True
        )
        
        # Vérifier que le fichier a été mis à jour
        wb = openpyxl.load_workbook(output_path)
        self.assertIn("Comptes 2023", wb.sheetnames)

    def test_ecrire_onglet_annuel_structure(self):
        """Test de la structure d'un onglet annuel."""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        write_annual_sheet(ws, self.recettes_test, self.depenses_test, 2023)
        
        # Vérifier quelques cellules clés
        self.assertIn("LIVRE DE COMPTES 2023", str(ws["A1"].value))
        
        # Chercher les sections recettes et dépenses
        found_recettes = False
        found_depenses = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "RECETTES" in str(cell.value):
                    found_recettes = True
                if cell.value and "DÉPENSES" in str(cell.value):
                    found_depenses = True
        
        self.assertTrue(found_recettes)
        self.assertTrue(found_depenses)


class TestIntegrationComplète(unittest.TestCase):
    """Tests d'intégration bout en bout."""

    def setUp(self):
        """Configuration des tests d'intégration."""
        self.temp_dir = tempfile.mkdtemp()
        self.temp_path = Path(self.temp_dir)

    def tearDown(self):
        """Nettoyage après tests."""
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    @patch('parser.camelot')
    def test_extraction_pdf_mock(self, mock_camelot):
        """Test d'extraction PDF avec des données mockées."""
        # Créer des données de test
        df_mock = pd.DataFrame([
            ["", "", "", "", "", ""],
            ["", "", "", "", "", ""],
            ["", "", "Libellé des opérations", "", "Débit", "Crédit"],
            ["opé.", "valeur", "", "", "", ""],
            ["26.01", "26.01", "Chèque test", "", "100,50", ""],
            ["27.01", "27.01", "Versement test", "", "", "200,75"],
        ])
        
        # Configurer le mock
        mock_table = Mock()
        mock_table.df = df_mock
        mock_camelot.read_pdf.return_value = [mock_table]
        
        # Tester l'extraction
        pdf_path = Path("test.pdf")
        result = extract_operations_pdf_camelot(pdf_path, 2023)
        
        # Vérifications
        self.assertEqual(len(result.income_transactions), 1)  # Versement
        self.assertEqual(len(result.expense_transactions), 1)  # Chèque
        self.assertEqual(result.income_transactions[0].montant, 200.75)  # Montant versement
        self.assertEqual(result.expense_transactions[0].montant, 100.50)  # Montant chèque

    def test_workflow_complet_mock(self):
        """Test du workflow complet avec données mockées."""
        with patch('parser.camelot') as mock_camelot:
            # Préparer les données mockées
            df_mock = pd.DataFrame([
                ["", "", "Libellé des opérations", "", "Débit", "Crédit"],
                ["", "", "", "", "", ""],
                ["01.01", "01.01", "Cotisation test", "", "", "150,00"],
                ["02.01", "02.01", "Achat matériel", "", "75,25", ""],
            ])
            
            mock_table = Mock()
            mock_table.df = df_mock
            mock_camelot.read_pdf.return_value = [mock_table]
            
            # Créer un fichier PDF fictif
            pdf_path = self.temp_path / "test.pdf"
            pdf_path.touch()
            
            # Test du workflow
            result = extract_operations_pdf_camelot(pdf_path, 2023)
            
            # Générer le fichier Excel
            output_path = self.temp_path / "workflow_test.xlsx"
            generate_or_update_excel(result.income_transactions, result.expense_transactions, output_path, 2023, force=True)
            
            # Vérifications finales
            self.assertTrue(output_path.exists())
            wb = openpyxl.load_workbook(output_path)
            self.assertIn("Comptes 2023", wb.sheetnames)
            self.assertIn("📈 Synthèse", wb.sheetnames)


class TestFonctionsStyle(unittest.TestCase):
    """Tests pour les fonctions de style Excel."""

    def test_style_bordure(self):
        """Test de création des bordures."""
        border_normale = style_border()
        border_fine = style_border(thin=True)
        
        self.assertEqual(border_normale.left.style, "hair")  # Default when thin=False gives "hair"
        self.assertEqual(border_fine.left.style, "thin")    # When thin=True gives "thin"

    def test_fill_couleur(self):
        """Test de création des remplissages de couleur."""
        fill_test = fill("FF0000")  # Rouge
        # OpenPyXL peut préfixer avec "00" pour l'alpha channel
        self.assertTrue(fill_test.fgColor.rgb in ["FF0000", "00FF0000"])

    def test_police_configuration(self):
        """Test de configuration des polices."""
        police_normale = font()
        police_gras = font(bold=True, size=14, color="FF0000")
        
        self.assertFalse(police_normale.bold)
        self.assertTrue(police_gras.bold)
        self.assertEqual(police_gras.size, 14)
        # OpenPyXL peut préfixer avec "00" pour l'alpha channel
        self.assertTrue(police_gras.color.rgb in ["FF0000", "00FF0000"])

    def test_centrer_alignement(self):
        """Test des alignements."""
        align_centre = center()
        align_gauche = center(horizontal="left")
        
        self.assertEqual(align_centre.horizontal, "center")
        self.assertEqual(align_gauche.horizontal, "left")


def run_tests():
    """Fonction pour exécuter tous les tests."""
    print("🧪 TESTS UNITAIRES")
    print("=" * 50)
    
    # Créer la suite de tests
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()
    
    # Ajouter toutes les classes de tests
    test_classes = [
        TestFonctionsUtilitaires,
        TestExtractionPDF, 
        TestGenerationExcel,
        TestIntegrationComplète,
        TestFonctionsStyle,
    ]
    
    for test_class in test_classes:
        tests = loader.loadTestsFromTestCase(test_class)
        suite.addTests(tests)
    
    # Exécuter les tests
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)
    
    # Afficher le résumé
    print(f"\n🎯 RÉSUMÉ DES TESTS UNITAIRES")
    print("=" * 50)
    print(f"✅ Tests réussis: {result.testsRun - len(result.failures) - len(result.errors)}")
    print(f"❌ Tests échoués: {len(result.failures)}")
    print(f"💥 Erreurs: {len(result.errors)}")
    print(f"📊 Total: {result.testsRun}")
    
    if result.failures:
        print("\n❌ ÉCHECS:")
        for test, traceback in result.failures:
            print(f"  - {test}: {traceback.split('AssertionError:')[-1].strip()}")
    
    if result.errors:
        print("\n💥 ERREURS:")
        for test, traceback in result.errors:
            print(f"  - {test}: {traceback.split('Error:')[-1].strip()}")
    
    return result.wasSuccessful()


if __name__ == "__main__":
    success = run_tests()
    exit(0 if success else 1)